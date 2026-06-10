#!/usr/bin/env python3
"""Analyze bid project Excel files and generate a customer-readable workbook."""

from __future__ import annotations

import argparse
import datetime as dt
import math
import os
import re
import sys
import threading
import traceback
import tkinter as tk
from dataclasses import dataclass, field
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover - shown in GUI/CLI runtime.
    raise SystemExit(f"缺少 Excel 依赖 openpyxl，请先安装：pip install openpyxl\n原因：{exc}")


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
BUCKET_ORDER = ["小于1000万", "1000万-2000万", "2000万-5000万", "5000万-1亿", "1亿及以上", "控制价缺失"]


@dataclass
class BidRow:
    company: str
    quote: float | None
    down_rate: float | None
    duration: str = ""
    quote_text: str = ""


@dataclass
class ProjectRecord:
    source_file: str
    sheet_name: str
    project_name: str = ""
    announcement_time: str = ""
    region: str = ""
    opening_count_raw: int | None = None
    evaluation_method: str = ""
    bid_type: str = ""
    control_price: float | None = None
    winning_price: float | None = None
    winner_company: str = ""
    bid_rows: list[BidRow] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)

    @property
    def quote_count(self) -> int:
        return len([row for row in self.bid_rows if row.quote is not None])

    @property
    def participant_count(self) -> int:
        return self.quote_count or int(self.opening_count_raw or 0)

    @property
    def bucket(self) -> str:
        return control_bucket(self.control_price)

    @property
    def winning_down_rate(self) -> float | None:
        return calc_down_rate(self.control_price, self.winning_price)

    @property
    def quote_values(self) -> list[float]:
        return [float(row.quote) for row in self.bid_rows if row.quote is not None]

    @property
    def quote_down_rates(self) -> list[float]:
        return [float(row.down_rate) for row in self.bid_rows if row.down_rate is not None]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u3000", " ")).strip()


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value) if isinstance(value, float) else False:
            return None
        return float(value)
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(",", "").replace("￥", "").replace("¥", "")
    multiplier = 1.0
    if "万元" in text:
        multiplier = 10000.0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0)) * multiplier


def parse_int(value: Any) -> int | None:
    number = parse_number(value)
    if number is None:
        return None
    return int(round(number))


def parse_rate(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        return number * 100.0 if abs(number) <= 1 else number
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    if not match:
        return None
    number = float(match.group(0))
    if "%" not in text and abs(number) <= 1:
        number *= 100.0
    return number


def calc_down_rate(control_price: float | None, amount: float | None) -> float | None:
    if control_price is None or amount is None or control_price <= 0:
        return None
    return (control_price - amount) / control_price * 100.0


def infer_control_price_from_quotes(bid_rows: list[BidRow]) -> float | None:
    inferred: list[float] = []
    for row in bid_rows:
        if row.quote is None or row.down_rate is None:
            continue
        denominator = 1.0 - row.down_rate / 100.0
        if denominator <= 0:
            continue
        value = row.quote / denominator
        if value > 0 and math.isfinite(value):
            inferred.append(value)
    if not inferred:
        return None
    inferred.sort()
    median = inferred[len(inferred) // 2]
    consistent = [value for value in inferred if abs(value - median) <= max(1.0, median * 0.001)]
    if len(consistent) < max(1, len(inferred) // 2):
        return None
    return sum(consistent) / len(consistent)


def control_bucket(control_price: float | None) -> str:
    if control_price is None:
        return "控制价缺失"
    if control_price < 10_000_000:
        return "小于1000万"
    if control_price < 20_000_000:
        return "1000万-2000万"
    if control_price < 50_000_000:
        return "2000万-5000万"
    if control_price < 100_000_000:
        return "5000万-1亿"
    return "1亿及以上"


def stats(values: list[float | int | None]) -> dict[str, Any]:
    clean = [float(value) for value in values if value is not None]
    if not clean:
        return {"样本数": 0, "最低": None, "最高": None, "平均": None}
    return {
        "样本数": len(clean),
        "最低": min(clean),
        "最高": max(clean),
        "平均": sum(clean) / len(clean),
    }


def row_values(ws, row_index: int) -> list[str]:
    return [clean_text(ws.cell(row_index, col).value) for col in range(1, ws.max_column + 1)]


def row_has_tokens(values: list[str], tokens: tuple[str, ...]) -> bool:
    joined = "|".join(values)
    return all(token in joined for token in tokens)


def find_label_row(ws, tokens: tuple[str, ...], start: int = 1) -> int | None:
    for row_index in range(start, ws.max_row + 1):
        if row_has_tokens(row_values(ws, row_index), tokens):
            return row_index
    return None


def map_header_values(headers: list[str], values: list[Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        mapped[header] = values[index] if index < len(values) else None
    return mapped


def get_first(mapped: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        for key, value in mapped.items():
            if alias == key:
                return value
    for alias in aliases:
        for key, value in mapped.items():
            if alias in key:
                return value
    return None


def infer_winner(project: ProjectRecord) -> None:
    if project.winner_company or project.winning_price is None:
        return
    for row in project.bid_rows:
        if row.quote is None:
            continue
        if abs(row.quote - project.winning_price) <= max(1.0, project.winning_price * 0.000001):
            project.winner_company = row.company
            return


def validate_project(project: ProjectRecord) -> None:
    if not project.project_name:
        project.issues.append("缺项目名称")
    if project.control_price is None:
        project.issues.append("缺控制价")
    if project.winning_price is None:
        project.issues.append("缺中标价")
    if not project.bid_rows:
        project.issues.append("缺各单位报价")
    if project.opening_count_raw is not None and project.quote_count and abs(project.opening_count_raw - project.quote_count) > 1:
        project.issues.append(f"开标家数({project.opening_count_raw})与有效报价家数({project.quote_count})不一致")
    if project.winning_price is not None and project.control_price is not None and project.winning_price > project.control_price * 1.2:
        project.issues.append("中标价明显高于控制价，请核对")


def parse_project_sheet(path: Path, ws) -> ProjectRecord | None:
    project_header_row = find_label_row(ws, ("项目名称",), 1)
    summary_header_row = find_label_row(ws, ("控制价", "中标价"), 1)
    bid_header_row = find_label_row(ws, ("公司名称", "投标报价"), 1)
    if project_header_row is None and summary_header_row is None and bid_header_row is None:
        return None

    project = ProjectRecord(source_file=str(path), sheet_name=ws.title)

    if project_header_row is not None and project_header_row < ws.max_row:
        headers = row_values(ws, project_header_row)
        values = [ws.cell(project_header_row + 1, col).value for col in range(1, ws.max_column + 1)]
        mapped = map_header_values(headers, values)
        project.project_name = clean_text(get_first(mapped, ("项目名称", "工程名称"))) or path.stem
        project.announcement_time = clean_text(get_first(mapped, ("公告时间", "发布时间", "开标时间")))
        project.region = clean_text(get_first(mapped, ("项目区域", "地区", "区域")))
        project.opening_count_raw = parse_int(get_first(mapped, ("开标家数", "投标家数", "参与家数", "竞争家数")))
        project.evaluation_method = clean_text(get_first(mapped, ("评标办法", "评审办法")))
        project.bid_type = clean_text(get_first(mapped, ("招标类型", "项目类型", "类型")))

    if summary_header_row is not None and summary_header_row < ws.max_row:
        headers = row_values(ws, summary_header_row)
        values = [ws.cell(summary_header_row + 1, col).value for col in range(1, ws.max_column + 1)]
        mapped = map_header_values(headers, values)
        project.control_price = parse_number(get_first(mapped, ("控制价", "招标控制价", "最高限价")))
        project.winning_price = parse_number(get_first(mapped, ("中标价", "中标金额", "中标价格")))

    if bid_header_row is not None:
        headers = row_values(ws, bid_header_row)
        company_col = quote_col = rate_col = duration_col = quote_text_col = None
        for index, header in enumerate(headers, start=1):
            if company_col is None and ("公司名称" in header or "投标人" in header or "单位名称" in header):
                company_col = index
            if quote_col is None and ("投标报价" in header or "报价" == header or "投标价" in header):
                quote_col = index
            if rate_col is None and ("下浮" in header or "降幅" in header):
                rate_col = index
            if duration_col is None and "工期" in header:
                duration_col = index
            if quote_text_col is None and "报价文本" in header:
                quote_text_col = index
        for row_index in range(bid_header_row + 1, ws.max_row + 1):
            company = clean_text(ws.cell(row_index, company_col).value if company_col else "")
            quote = parse_number(ws.cell(row_index, quote_col).value if quote_col else None)
            if not company and quote is None:
                continue
            if not company and quote is not None:
                company = f"未识别单位{len(project.bid_rows) + 1}"
            down_rate = parse_rate(ws.cell(row_index, rate_col).value if rate_col else None)
            if down_rate is None:
                down_rate = calc_down_rate(project.control_price, quote)
            duration = clean_text(ws.cell(row_index, duration_col).value if duration_col else "")
            quote_text = clean_text(ws.cell(row_index, quote_text_col).value if quote_text_col else "")
            project.bid_rows.append(BidRow(company=company, quote=quote, down_rate=down_rate, duration=duration, quote_text=quote_text))

    if project.control_price is None:
        inferred_control = infer_control_price_from_quotes(project.bid_rows)
        if inferred_control is not None:
            project.control_price = inferred_control
    for row in project.bid_rows:
        if row.down_rate is None:
            row.down_rate = calc_down_rate(project.control_price, row.quote)

    project.bid_rows.sort(key=lambda item: (-999999.0 if item.down_rate is None else -item.down_rate, item.company))
    infer_winner(project)
    validate_project(project)
    return project


def parse_project_file(path: Path) -> list[ProjectRecord]:
    wb = load_workbook(path, data_only=True)
    projects: list[ProjectRecord] = []
    for ws in wb.worksheets:
        parsed = parse_project_sheet(path, ws)
        if parsed is not None:
            projects.append(parsed)
    return projects


def collect_input_files(paths: list[str]) -> list[Path]:
    def should_skip(item: Path) -> bool:
        return item.name.startswith("~$")

    files: list[Path] = []
    for raw_path in paths:
        path = Path(raw_path).expanduser()
        if path.is_dir():
            for item in sorted(path.rglob("*")):
                if should_skip(item):
                    continue
                if item.suffix.lower() in SUPPORTED_EXTENSIONS:
                    files.append(item)
        elif path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS and not should_skip(path):
            files.append(path)
    seen: set[str] = set()
    unique: list[Path] = []
    for file_path in files:
        key = str(file_path.resolve())
        if key in seen:
            continue
        seen.add(key)
        unique.append(file_path)
    return unique


def analyze_projects(projects: list[ProjectRecord]) -> dict[str, Any]:
    all_quote_rates = [rate for project in projects for rate in project.quote_down_rates]
    winning_rates = [project.winning_down_rate for project in projects if project.winning_down_rate is not None]
    participant_counts = [project.participant_count for project in projects if project.participant_count > 0]
    buckets: dict[str, dict[str, Any]] = {}
    for bucket in BUCKET_ORDER:
        bucket_projects = [project for project in projects if project.bucket == bucket]
        bucket_quote_rates = [rate for project in bucket_projects for rate in project.quote_down_rates]
        bucket_winning_rates = [project.winning_down_rate for project in bucket_projects if project.winning_down_rate is not None]
        bucket_participants = [project.participant_count for project in bucket_projects if project.participant_count > 0]
        buckets[bucket] = {
            "项目数": len(bucket_projects),
            "控制价统计": stats([project.control_price for project in bucket_projects]),
            "报价下浮率统计": stats(bucket_quote_rates),
            "中标下浮率统计": stats(bucket_winning_rates),
            "项目参与竞争家数统计": stats(bucket_participants),
        }
    return {
        "项目总数": len(projects),
        "完整项目数": sum(1 for project in projects if not project.issues),
        "报价明细总数": sum(project.quote_count for project in projects),
        "报价下浮率统计": stats(all_quote_rates),
        "中标下浮率统计": stats(winning_rates),
        "项目参与竞争家数统计": stats(participant_counts),
        "控制价档位统计": buckets,
    }


def money_fmt(value: float | None) -> float | None:
    return None if value is None else round(float(value), 2)


def rate_fmt(value: float | None) -> float | None:
    return None if value is None else round(float(value), 4)


def winning_rate_summary_fmt(analysis: dict[str, Any], key: str) -> float | str:
    winning_stats = analysis["中标下浮率统计"]
    if winning_stats["样本数"] == 0:
        return "缺中标价数据"
    return rate_fmt(winning_stats[key])


def set_common_style(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        max_len = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, column).value
            if value is not None:
                max_len = max(max_len, min(36, len(str(value)) + 2))
        ws.column_dimensions[letter].width = max_len


def append_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def create_output_workbook(projects: list[ProjectRecord], output_path: Path) -> None:
    analysis = analyze_projects(projects)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("汇总看板")
    append_rows(ws, [
        ["指标", "数值", "说明"],
        ["项目总数", analysis["项目总数"], "输入 Excel 中成功识别的项目数"],
        ["完整项目数", analysis["完整项目数"], "没有数据问题标记的项目"],
        ["报价明细总数", analysis["报价明细总数"], "所有项目的有效投标报价行数"],
        ["报价最高下浮率", rate_fmt(analysis["报价下浮率统计"]["最高"]), "所有投标报价下浮率中的最高值"],
        ["报价最低下浮率", rate_fmt(analysis["报价下浮率统计"]["最低"]), "所有投标报价下浮率中的最低值"],
        ["报价平均下浮率", rate_fmt(analysis["报价下浮率统计"]["平均"]), "所有投标报价下浮率的平均值"],
        ["中标最高下浮率", winning_rate_summary_fmt(analysis, "最高"), "所有项目中标价下浮率最高值；输入缺中标价时无法计算"],
        ["中标最低下浮率", winning_rate_summary_fmt(analysis, "最低"), "所有项目中标价下浮率最低值；输入缺中标价时无法计算"],
        ["中标平均下浮率", winning_rate_summary_fmt(analysis, "平均"), "所有项目中标价下浮率平均值；输入缺中标价时无法计算"],
        ["项目参与竞争家数最高", analysis["项目参与竞争家数统计"]["最高"], "按每个项目有效报价家数统计"],
        ["项目参与竞争家数最低", analysis["项目参与竞争家数统计"]["最低"], "按每个项目有效报价家数统计"],
        ["项目参与竞争家数平均", round(analysis["项目参与竞争家数统计"]["平均"], 2) if analysis["项目参与竞争家数统计"]["平均"] is not None else None, "按每个项目有效报价家数统计"],
    ])

    ws = wb.create_sheet("控制价档位对比")
    append_rows(ws, [[
        "控制价档位", "项目数", "控制价最低", "控制价最高", "控制价平均",
        "报价下浮率最低", "报价下浮率最高", "报价下浮率平均",
        "中标下浮率最低", "中标下浮率最高", "中标下浮率平均",
        "项目参与竞争家数最低", "项目参与竞争家数最高", "项目参与竞争家数平均",
    ]])
    for bucket in BUCKET_ORDER:
        item = analysis["控制价档位统计"][bucket]
        control_stats = item["控制价统计"]
        quote_stats = item["报价下浮率统计"]
        winning_stats = item["中标下浮率统计"]
        participant_stats = item["项目参与竞争家数统计"]
        ws.append([
            bucket, item["项目数"],
            money_fmt(control_stats["最低"]), money_fmt(control_stats["最高"]), money_fmt(control_stats["平均"]),
            rate_fmt(quote_stats["最低"]), rate_fmt(quote_stats["最高"]), rate_fmt(quote_stats["平均"]),
            rate_fmt(winning_stats["最低"]), rate_fmt(winning_stats["最高"]), rate_fmt(winning_stats["平均"]),
            participant_stats["最低"], participant_stats["最高"],
            round(participant_stats["平均"], 2) if participant_stats["平均"] is not None else None,
        ])

    ws = wb.create_sheet("项目汇总")
    append_rows(ws, [[
        "项目名称", "公告时间", "项目区域", "控制价档位", "控制价", "中标价", "中标单位", "中标下浮率",
        "开标家数(原表)", "有效报价家数", "项目参与竞争家数", "评标办法", "招标类型",
        "最高报价", "最低报价", "平均报价", "最高下浮率", "最低下浮率", "平均下浮率", "数据问题", "来源文件",
    ]])
    for project in projects:
        quotes = project.quote_values
        rates = project.quote_down_rates
        ws.append([
            project.project_name, project.announcement_time, project.region, project.bucket,
            money_fmt(project.control_price), money_fmt(project.winning_price), project.winner_company,
            rate_fmt(project.winning_down_rate),
            project.opening_count_raw, project.quote_count, project.participant_count,
            project.evaluation_method, project.bid_type,
            money_fmt(max(quotes) if quotes else None), money_fmt(min(quotes) if quotes else None),
            money_fmt(sum(quotes) / len(quotes) if quotes else None),
            rate_fmt(max(rates) if rates else None), rate_fmt(min(rates) if rates else None),
            rate_fmt(sum(rates) / len(rates) if rates else None),
            "；".join(project.issues), project.source_file,
        ])

    ws = wb.create_sheet("报价明细")
    append_rows(ws, [["项目名称", "控制价档位", "控制价", "公司名称", "投标报价", "报价下浮率", "是否中标单位", "工期", "报价文本", "来源文件"]])
    for project in projects:
        for row in project.bid_rows:
            is_winner = "是" if project.winner_company and row.company == project.winner_company else ""
            ws.append([
                project.project_name, project.bucket, money_fmt(project.control_price), row.company,
                money_fmt(row.quote), rate_fmt(row.down_rate), is_winner, row.duration, row.quote_text, project.source_file,
            ])

    ws = wb.create_sheet("中标下浮率")
    append_rows(ws, [["项目名称", "控制价档位", "控制价", "中标价", "中标单位", "中标下浮率", "项目参与竞争家数", "来源文件"]])
    for project in sorted(projects, key=lambda item: -999999.0 if item.winning_down_rate is None else -item.winning_down_rate):
        ws.append([
            project.project_name, project.bucket, money_fmt(project.control_price), money_fmt(project.winning_price),
            project.winner_company, rate_fmt(project.winning_down_rate), project.participant_count, project.source_file,
        ])

    ws = wb.create_sheet("数据问题")
    append_rows(ws, [["项目名称", "问题", "来源文件", "Sheet"]])
    for project in projects:
        if not project.issues:
            continue
        for issue in project.issues:
            ws.append([project.project_name, issue, project.source_file, project.sheet_name])

    for sheet in wb.worksheets:
        set_common_style(sheet)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    header = clean_text(sheet.cell(1, cell.column).value)
                    if "下浮率" in header:
                        cell.number_format = "0.0000"
                    elif "价" in header or "报价" in header:
                        cell.number_format = "#,##0.00"
                    elif "家数" in header or "项目数" in header:
                        cell.number_format = "0"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def default_output_path() -> Path:
    desktop = Path.home() / "Desktop"
    base = desktop if desktop.exists() else Path.cwd()
    return base / f"项目报价分析_{dt.datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"


def timestamped_output_path(path: Path) -> Path:
    timestamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    if re.search(r"_\d{8}-\d{6}$", path.stem):
        candidate = path
    else:
        candidate = path.with_name(f"{path.stem}_{timestamp}{path.suffix or '.xlsx'}")
    if not candidate.exists():
        return candidate
    index = 2
    while True:
        next_candidate = candidate.with_name(f"{candidate.stem}_{index}{candidate.suffix}")
        if not next_candidate.exists():
            return next_candidate
        index += 1


def run_analysis(input_paths: list[str], output_path: str | None = None) -> tuple[Path, list[ProjectRecord]]:
    files = collect_input_files(input_paths)
    if not files:
        raise RuntimeError("没有找到可分析的 .xlsx/.xlsm 文件。")
    projects: list[ProjectRecord] = []
    parse_errors: list[str] = []
    for file_path in files:
        try:
            parsed = parse_project_file(file_path)
            if parsed:
                projects.extend(parsed)
            else:
                parse_errors.append(f"{file_path}: 未识别到项目数据")
        except Exception as exc:
            parse_errors.append(f"{file_path}: {exc}")
    if not projects:
        raise RuntimeError("没有成功识别任何项目。\n" + "\n".join(parse_errors[:20]))
    for error in parse_errors:
        projects.append(ProjectRecord(source_file=error, sheet_name="", project_name="未识别文件", issues=[error]))
    output = timestamped_output_path(Path(output_path).expanduser()) if output_path else default_output_path()
    create_output_workbook(projects, output)
    return output, projects


class AnalyzerApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("项目报价分析器")
        self.root.geometry("980x680")
        self.input_paths: list[str] = []
        self.output_var = tk.StringVar(value=str(default_output_path()))
        self.running = False
        self._build_ui()

    def _build_ui(self) -> None:
        top = ttk.Frame(self.root, padding=14)
        top.pack(fill="x")
        ttk.Button(top, text="添加Excel", command=self.add_files).pack(side="left")
        ttk.Button(top, text="添加文件夹", command=self.add_folder).pack(side="left", padx=8)
        ttk.Button(top, text="清空", command=self.clear_files).pack(side="left")
        ttk.Button(top, text="开始分析", command=self.start_analysis).pack(side="right")

        output_frame = ttk.Frame(self.root, padding=(14, 0, 14, 8))
        output_frame.pack(fill="x")
        ttk.Label(output_frame, text="输出Excel").pack(side="left")
        ttk.Entry(output_frame, textvariable=self.output_var).pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(output_frame, text="选择", command=self.choose_output).pack(side="right")

        file_frame = ttk.Labelframe(self.root, text="输入项目Excel/文件夹", padding=8)
        file_frame.pack(fill="both", expand=True, padx=14, pady=(0, 8))
        self.file_list = tk.Listbox(file_frame, height=12)
        self.file_list.pack(fill="both", expand=True)

        log_frame = ttk.Labelframe(self.root, text="运行日志", padding=8)
        log_frame.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.log = tk.Text(log_frame, height=12, wrap="word")
        self.log.pack(fill="both", expand=True)
        self.append_log("说明：支持一次选择多个项目 Excel，或选择一个包含多个项目 Excel 的文件夹。\n")

    def append_log(self, text: str) -> None:
        self.log.insert("end", text)
        self.log.see("end")

    def refresh_list(self) -> None:
        self.file_list.delete(0, "end")
        for path in self.input_paths:
            self.file_list.insert("end", path)

    def add_files(self) -> None:
        paths = filedialog.askopenfilenames(title="选择项目Excel", filetypes=[("Excel", "*.xlsx *.xlsm"), ("All files", "*.*")])
        for path in paths:
            if path not in self.input_paths:
                self.input_paths.append(path)
        self.refresh_list()

    def add_folder(self) -> None:
        path = filedialog.askdirectory(title="选择包含项目Excel的文件夹")
        if path and path not in self.input_paths:
            self.input_paths.append(path)
        self.refresh_list()

    def clear_files(self) -> None:
        self.input_paths.clear()
        self.refresh_list()

    def choose_output(self) -> None:
        path = filedialog.asksaveasfilename(title="选择输出Excel", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if path:
            self.output_var.set(path)

    def start_analysis(self) -> None:
        if self.running:
            messagebox.showinfo("提示", "正在分析中。")
            return
        if not self.input_paths:
            messagebox.showwarning("提示", "请先添加 Excel 文件或文件夹。")
            return
        self.running = True
        self.append_log("开始分析...\n")

        def worker() -> None:
            try:
                output, projects = run_analysis(self.input_paths, self.output_var.get().strip())
                self.root.after(0, self.analysis_done, output, projects)
            except Exception as exc:
                detail = traceback.format_exc()
                self.root.after(0, self.analysis_failed, exc, detail)

        threading.Thread(target=worker, daemon=True).start()

    def analysis_done(self, output: Path, projects: list[ProjectRecord]) -> None:
        self.running = False
        self.append_log(f"完成：识别项目 {len(projects)} 个。\n输出：{output}\n")
        self.output_var.set(str(default_output_path()))
        messagebox.showinfo("完成", f"分析完成。\n识别项目：{len(projects)} 个\n输出：{output}")

    def analysis_failed(self, exc: Exception, detail: str) -> None:
        self.running = False
        self.append_log(f"失败：{exc}\n{detail}\n")
        messagebox.showerror("分析失败", str(exc))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="批量分析项目报价 Excel")
    parser.add_argument("inputs", nargs="*", help="项目 Excel 文件或文件夹")
    parser.add_argument("-o", "--output", default="", help="输出分析 Excel 路径")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.inputs:
        output, projects = run_analysis(args.inputs, args.output or None)
        print(f"完成：识别项目 {len(projects)} 个")
        print(f"输出：{output}")
        return
    root = tk.Tk()
    AnalyzerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
